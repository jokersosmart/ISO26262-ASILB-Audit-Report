#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
md_to_excel.py — Reversible Markdown → Excel Converter

This script reads the DFA_Report_Reversible.md file and reconstructs
an Excel workbook that closely approximates the original Excel file.

Usage:
    python3 md_to_excel.py <input_md> <output_xlsx>

Example:
    python3 md_to_excel.py ../markdown/DFA_Report_Reversible.md ../output_reconstructed.xlsx
"""

import re
import json
import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import column_index_from_string

def parse_markdown(md_path):
    """Parse the reversible Markdown file into a structured dictionary."""
    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()

    sheets = {}
    # Split by sheet headers
    sheet_sections = re.split(r'^---\s*\n\n## Sheet: (.+)$', content, flags=re.MULTILINE)

    for i in range(1, len(sheet_sections), 2):
        sheet_name = sheet_sections[i].strip()
        sheet_content = sheet_sections[i + 1] if i + 1 < len(sheet_sections) else ''
        sheets[sheet_name] = parse_sheet_section(sheet_content)

    return sheets

def parse_sheet_section(section):
    """Parse a single sheet section into structured data."""
    data = {
        'state': 'visible',
        'merged_cells': [],
        'col_dims': {},
        'row_dims': {},
        'cells': [],
        'images': []
    }

    # Extract state
    state_match = re.search(r'\*\*State\*\*: `(\w+)`', section)
    if state_match:
        data['state'] = state_match.group(1)

    # Extract merged cells
    merged_match = re.search(r'\*\*Merged Cells\*\*: `(.+?)`', section)
    if merged_match:
        data['merged_cells'] = [mc.strip() for mc in merged_match.group(1).split(',')]

    # Extract column dimensions
    col_dims_match = re.search(r'### Column Dimensions\s*\n```json\n(.*?)\n```', section, re.DOTALL)
    if col_dims_match:
        try:
            data['col_dims'] = json.loads(col_dims_match.group(1))
        except json.JSONDecodeError:
            pass

    # Extract row dimensions
    row_dims_match = re.search(r'### Row Dimensions\s*\n```json\n(.*?)\n```', section, re.DOTALL)
    if row_dims_match:
        try:
            data['row_dims'] = json.loads(row_dims_match.group(1))
        except json.JSONDecodeError:
            pass

    # Extract cell data from table
    cell_table_match = re.search(
        r'### Cell Data.*?\n\|.*?\n\|.*?\n(.*?)(?=\n### |\n---|\Z)',
        section, re.DOTALL
    )
    if cell_table_match:
        for line in cell_table_match.group(1).strip().split('\n'):
            if line.startswith('|') and not line.startswith('|:'):
                parts = [p.strip() for p in line.split('|')[1:-1]]
                if len(parts) >= 4:
                    cell_ref = parts[0]
                    value = parts[1].replace('<br>', '\n').replace('\\|', '|')
                    formula = parts[2].strip('`') if parts[2] else ''
                    style_str = parts[3].strip('`') if parts[3] else ''
                    try:
                        style = json.loads(style_str) if style_str else {}
                    except json.JSONDecodeError:
                        style = {}
                    data['cells'].append({
                        'ref': cell_ref,
                        'value': value,
                        'formula': formula,
                        'style': style
                    })

    return data

def apply_style(cell, style_dict):
    """Apply style dictionary to an openpyxl cell."""
    if not style_dict:
        return

    if 'font' in style_dict:
        f = style_dict['font']
        cell.font = Font(
            name=f.get('name', 'Calibri'),
            size=f.get('size', 11),
            bold=f.get('bold', False),
            italic=f.get('italic', False),
            color=f.get('color') if f.get('color') and not f['color'].startswith('Values') else None
        )

    if 'fill' in style_dict:
        fl = style_dict['fill']
        fg = fl.get('fgColor')
        if fg and not str(fg).startswith('Values'):
            cell.fill = PatternFill(
                patternType=fl.get('pattern', 'solid'),
                fgColor=fg
            )

    if 'border' in style_dict:
        b = style_dict['border']
        cell.border = Border(
            left=Side(style=b.get('left')) if b.get('left') else Side(),
            right=Side(style=b.get('right')) if b.get('right') else Side(),
            top=Side(style=b.get('top')) if b.get('top') else Side(),
            bottom=Side(style=b.get('bottom')) if b.get('bottom') else Side()
        )

    if 'alignment' in style_dict:
        a = style_dict['alignment']
        cell.alignment = Alignment(
            horizontal=a.get('horizontal'),
            vertical=a.get('vertical'),
            wrap_text=a.get('wrap_text', False)
        )

def reconstruct_excel(sheets, output_path):
    """Reconstruct an Excel workbook from parsed Markdown data."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, sheet_data in sheets.items():
        ws = wb.create_sheet(title=sheet_name)

        # Set sheet state
        if sheet_data['state'] == 'hidden':
            ws.sheet_state = 'hidden'

        # Set column dimensions
        for col_letter, dims in sheet_data.get('col_dims', {}).items():
            if dims.get('width') and dims['width'] != 'default':
                ws.column_dimensions[col_letter].width = float(dims['width'])
            if dims.get('hidden'):
                ws.column_dimensions[col_letter].hidden = True

        # Set row dimensions
        for row_num, dims in sheet_data.get('row_dims', {}).items():
            row_int = int(row_num)
            if dims.get('height') and dims['height'] != 'default':
                ws.row_dimensions[row_int].height = float(dims['height'])
            if dims.get('hidden'):
                ws.row_dimensions[row_int].hidden = True

        # Set merged cells
        for mc_range in sheet_data.get('merged_cells', []):
            try:
                ws.merge_cells(mc_range)
            except Exception:
                pass

        # Set cell data
        for cell_data in sheet_data.get('cells', []):
            cell = ws[cell_data['ref']]
            if cell_data['formula']:
                cell.value = cell_data['formula']
            elif cell_data['value']:
                # Try to convert to number
                try:
                    cell.value = float(cell_data['value'])
                except ValueError:
                    cell.value = cell_data['value']
            apply_style(cell, cell_data['style'])

    wb.save(output_path)
    print(f"Reconstructed Excel saved to: {output_path}")

def main():
    if len(sys.argv) < 3:
        print("Usage: python3 md_to_excel.py <input_md> <output_xlsx>")
        sys.exit(1)

    md_path = sys.argv[1]
    output_path = sys.argv[2]

    print(f"Parsing Markdown: {md_path}")
    sheets = parse_markdown(md_path)
    print(f"Found {len(sheets)} sheets")

    print(f"Reconstructing Excel...")
    reconstruct_excel(sheets, output_path)

if __name__ == "__main__":
    main()
