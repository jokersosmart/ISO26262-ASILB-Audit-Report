#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import json
from openpyxl.utils import get_column_letter
import os

# --- Configuration ---
EXCEL_FILE_PATH = '/home/ubuntu/upload/RD-03-009-01.ADFA報告DFAReport.xlsx'
OUTPUT_MD_PATH = '/home/ubuntu/DFA_Report_Reversible.md'
IMAGE_DIR = '/home/ubuntu/extracted_images'

# --- Helper Functions ---
def get_cell_style(cell):
    """Extracts a dictionary of style attributes from a cell."""
    style = {}
    if cell.has_style:
        if cell.font and (cell.font.name or cell.font.sz or cell.font.b or cell.font.i or cell.font.color):
            font_style = {}
            if cell.font.name: font_style['name'] = cell.font.name
            if cell.font.sz: font_style['size'] = cell.font.sz
            if cell.font.b: font_style['bold'] = True
            if cell.font.i: font_style['italic'] = True
            if cell.font.color and cell.font.color.rgb: font_style['color'] = str(cell.font.color.rgb)
            if font_style: style['font'] = font_style

        if cell.fill and cell.fill.patternType is not None:
            fill_style = {'pattern': cell.fill.patternType}
            if cell.fill.fgColor and cell.fill.fgColor.rgb: fill_style['fgColor'] = str(cell.fill.fgColor.rgb)
            if fill_style: style['fill'] = fill_style

        if cell.border and (cell.border.left or cell.border.right or cell.border.top or cell.border.bottom):
            border_style = {}
            if cell.border.left and cell.border.left.style: border_style['left'] = cell.border.left.style
            if cell.border.right and cell.border.right.style: border_style['right'] = cell.border.right.style
            if cell.border.top and cell.border.top.style: border_style['top'] = cell.border.top.style
            if cell.border.bottom and cell.border.bottom.style: border_style['bottom'] = cell.border.bottom.style
            if border_style: style['border'] = border_style

        if cell.alignment and (cell.alignment.horizontal or cell.alignment.vertical or cell.alignment.wrap_text):
            align_style = {}
            if cell.alignment.horizontal: align_style['horizontal'] = cell.alignment.horizontal
            if cell.alignment.vertical: align_style['vertical'] = cell.alignment.vertical
            if cell.alignment.wrap_text: align_style['wrap_text'] = True
            if align_style: style['alignment'] = align_style
            
    return style

def main():
    """Main function to convert Excel to reversible Markdown."""
    try:
        # Load workbook for formulas and styles
        wb_formulas = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=False)
        # Load workbook for calculated values
        wb_values = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    except FileNotFoundError:
        print(f"Error: Excel file not found at {EXCEL_FILE_PATH}")
        return

    with open(OUTPUT_MD_PATH, 'w', encoding='utf-8') as md_file:
        md_file.write("# ADFA/DFA Report - Reversible Markdown Format\n\n")
        md_file.write("This document represents the structure, data, formulas, and formatting of the original Excel file in a reversible Markdown format.\n\n")

        for sheet_name in wb_formulas.sheetnames:
            sheet_formulas = wb_formulas[sheet_name]
            sheet_values = wb_values[sheet_name]
            md_file.write(f"---\n\n## Sheet: {sheet_name}\n\n")
            md_file.write(f"- **State**: `{sheet_formulas.sheet_state}`\n")

            # --- Merged Cells ---
            merged_ranges = [str(mc_range) for mc_range in sheet_formulas.merged_cells.ranges]
            if merged_ranges:
                md_file.write(f"- **Merged Cells**: `{', '.join(merged_ranges)}`\n")

            # --- Column Dimensions ---
            md_file.write("\n### Column Dimensions\n")
            col_dims = {}
            for i in range(1, sheet_formulas.max_column + 1):
                col_letter = get_column_letter(i)
                dim = sheet_formulas.column_dimensions[col_letter]
                if dim.width != sheet_formulas.sheet_format.baseColWidth or dim.hidden:
                    col_dims[col_letter] = {'width': f"{dim.width:.2f}" if dim.width else 'default', 'hidden': dim.hidden}
            if col_dims:
                md_file.write("```json\n" + json.dumps(col_dims, indent=2) + "\n```\n")
            else:
                md_file.write("All columns have default width and are visible.\n")

            # --- Row Dimensions ---
            md_file.write("\n### Row Dimensions\n")
            row_dims = {}
            for i in range(1, sheet_formulas.max_row + 1):
                dim = sheet_formulas.row_dimensions[i]
                if (dim.height is not None and dim.height != sheet_formulas.sheet_format.defaultRowHeight) or dim.hidden:
                    row_dims[i] = {'height': f"{dim.height:.2f}" if dim.height is not None else 'default', 'hidden': dim.hidden}
            if row_dims:
                md_file.write("```json\n" + json.dumps(row_dims, indent=2) + "\n```\n")
            else:
                md_file.write("All rows have default height and are visible.\n")

            # --- Cell Data and Formulas ---
            md_file.write("\n### Cell Data, Formulas, and Styles\n")
            md_file.write("| Cell | Value (Displayed) | Formula | Style (JSON) |\n")
            md_file.write("|:-----|:------------------|:--------|:-------------|\n")

            for row in sheet_formulas.iter_rows():
                for cell in row:
                    style = get_cell_style(cell)
                    if cell.value is not None or style:
                        # Get displayed value from the data_only workbook
                        value_cell = sheet_values[cell.coordinate]
                        value_str = str(value_cell.value).replace('\n', '<br>').replace('|', '\\|') if value_cell.value is not None else ''
                        
                        # Get formula from the main workbook
                        formula_str = f"`{cell.value}`" if cell.data_type == 'f' else ''
                        
                        style_str = f"`{json.dumps(style)}`" if style else ''
                        md_file.write(f"| {cell.coordinate} | {value_str} | {formula_str} | {style_str} |\n")

            # --- Images ---
            if sheet_formulas._images:
                md_file.write("\n### Images\n")
                for i, img in enumerate(sheet_formulas._images):
                    img_filename = f"{sheet_name.replace(' ', '_')}_img{i}.{img.format}"
                    img_path = os.path.join(IMAGE_DIR, img_filename)
                    md_file.write(f"- **Image {i}**:\n")
                    md_file.write(f"  - **Source Path**: `{img_path}`\n")
                    md_file.write(f"  - **Anchor**: `{img.anchor._from.col + 1},{img.anchor._from.row + 1}`\n")
                    md_file.write(f"  - **Size (pixels)**: {img.width}x{img.height}\n")

            md_file.write("\n")

    print(f"Successfully converted Excel to reversible Markdown at {OUTPUT_MD_PATH}")

if __name__ == "__main__":
    main()
