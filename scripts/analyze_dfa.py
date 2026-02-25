import openpyxl
from openpyxl.utils import get_column_letter
import json
import os

EXCEL_PATH = "/home/ubuntu/upload/RD-03-009-01.ADFA報告DFAReport.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
wb_data = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

output = {}

print("=" * 80)
print(f"Workbook: {os.path.basename(EXCEL_PATH)}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws_data = wb_data[sheet_name]
    
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    print(f"  Sheet state: {ws.sheet_state}")
    
    # Merged cells
    merged = list(ws.merged_cells.ranges)
    if merged:
        print(f"  Merged cells ({len(merged)}):")
        for m in merged:
            print(f"    {m}")
    
    # Hidden rows
    hidden_rows = []
    for row in range(1, ws.max_row + 1):
        rd = ws.row_dimensions.get(row)
        if rd and rd.hidden:
            hidden_rows.append(row)
    if hidden_rows:
        print(f"  Hidden rows: {hidden_rows}")
    
    # Hidden columns
    hidden_cols = []
    for col in range(1, ws.max_column + 1):
        cd = ws.column_dimensions.get(get_column_letter(col))
        if cd and cd.hidden:
            hidden_cols.append(get_column_letter(col))
    if hidden_cols:
        print(f"  Hidden columns: {hidden_cols}")
    
    # Formulas
    formulas = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'computed_value': ws_data[cell.coordinate].value
                })
    if formulas:
        print(f"  Formulas ({len(formulas)}):")
        for f in formulas:
            print(f"    {f['cell']}: {f['formula']}  => {f['computed_value']}")
    
    # Images
    images = ws._images
    if images:
        print(f"  Images ({len(images)}):")
        for idx, img in enumerate(images):
            try:
                w = img.width
                h = img.height
                print(f"    Image {idx}: size=({w}x{h})")
            except:
                print(f"    Image {idx}: (could not read details)")
    
    # Charts
    charts = ws._charts
    if charts:
        print(f"  Charts ({len(charts)}):")
        for idx, chart in enumerate(charts):
            print(f"    Chart {idx}: type={type(chart).__name__}, title={chart.title}")
    
    # Data validation
    if ws.data_validations and ws.data_validations.dataValidation:
        print(f"  Data Validations ({len(ws.data_validations.dataValidation)}):")
        for dv in ws.data_validations.dataValidation:
            print(f"    {dv.sqref}: type={dv.type}, formula1={dv.formula1}")
    
    # Conditional formatting
    if ws.conditional_formatting:
        print(f"  Conditional Formatting ({len(ws.conditional_formatting)}):")
        for cf in ws.conditional_formatting:
            print(f"    {cf}")
    
    # Print all cell data (first 100 rows max for overview)
    print(f"\n  --- Cell Data (up to 100 rows) ---")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), min_col=1, max_col=ws.max_column, values_only=False):
        row_data = []
        for cell in row:
            if cell.value is not None:
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    computed = ws_data[cell.coordinate].value
                    row_data.append(f"{cell.coordinate}=[F]{val}=>{computed}")
                else:
                    row_data.append(f"{cell.coordinate}={val}")
        if row_data:
            print(f"    Row {row[0].row}: {' | '.join(row_data)}")

# Also check named ranges
print(f"\n{'='*80}")
print("Named Ranges / Defined Names:")
for dn in wb.defined_names.definedName:
    print(f"  {dn.name} = {dn.attr_text}")

wb.close()
wb_data.close()
