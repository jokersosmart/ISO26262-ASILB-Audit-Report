#!/usr/bin/env python3
"""
全面分析 RD-03-008-01FMEDAReport.xlsx 的結構：
- 所有 sheet 名稱、可見/隱藏狀態
- 每個 sheet 的行列數、合併儲存格、隱藏行列
- 所有公式（含跨頁引用）
- 所有圖片/嵌入物件
- 資料內容摘要
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json, os

EXCEL_PATH = "/home/ubuntu/upload/RD-03-008-01FMEDAReport.xlsx"
OUTPUT_PATH = "/home/ubuntu/fmeda_analysis_output.txt"

# Load with data_only=False to see formulas
wb_formula = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
# Load with data_only=True to see computed values
wb_data = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

out = []
def p(text=""):
    out.append(str(text))

p("=" * 80)
p("FMEDA REPORT EXCEL 完整結構分析")
p(f"檔案: {EXCEL_PATH}")
p(f"Sheet 數量: {len(wb_formula.sheetnames)}")
p("=" * 80)

# List all sheets with visibility
p("\n## 所有 Sheet 列表")
for i, name in enumerate(wb_formula.sheetnames):
    ws = wb_formula[name]
    visibility = ws.sheet_state  # 'visible', 'hidden', 'veryHidden'
    p(f"  [{i+1}] {name} — 狀態: {visibility}")

# Analyze each sheet
for sheet_name in wb_formula.sheetnames:
    ws_f = wb_formula[sheet_name]
    ws_d = wb_data[sheet_name]
    
    p("\n" + "=" * 80)
    p(f"SHEET: {sheet_name}")
    p(f"  可見性: {ws_f.sheet_state}")
    p(f"  行數: {ws_f.max_row}, 列數: {ws_f.max_column}")
    p(f"  合併儲存格: {len(ws_f.merged_cells.ranges)}")
    
    # Merged cells detail
    if ws_f.merged_cells.ranges:
        p(f"\n  ### 合併儲存格清單:")
        for mc in sorted(ws_f.merged_cells.ranges, key=str):
            p(f"    {mc}")
    
    # Hidden rows
    hidden_rows = []
    for row_idx in range(1, ws_f.max_row + 1):
        rd = ws_f.row_dimensions.get(row_idx)
        if rd and rd.hidden:
            hidden_rows.append(row_idx)
    if hidden_rows:
        p(f"\n  ### 隱藏行: {hidden_rows}")
    
    # Hidden columns
    hidden_cols = []
    for col_idx in range(1, ws_f.max_column + 1):
        col_letter = get_column_letter(col_idx)
        cd = ws_f.column_dimensions.get(col_letter)
        if cd and cd.hidden:
            hidden_cols.append(col_letter)
    if hidden_cols:
        p(f"\n  ### 隱藏列: {hidden_cols}")
    
    # Column widths
    p(f"\n  ### 列寬:")
    for col_idx in range(1, min(ws_f.max_column + 1, 50)):
        col_letter = get_column_letter(col_idx)
        cd = ws_f.column_dimensions.get(col_letter)
        width = cd.width if cd and cd.width else "default"
        p(f"    {col_letter}: {width}")
    
    # Row heights
    p(f"\n  ### 行高 (非預設):")
    for row_idx in range(1, min(ws_f.max_row + 1, 200)):
        rd = ws_f.row_dimensions.get(row_idx)
        if rd and rd.height and rd.height != 15:
            p(f"    Row {row_idx}: {rd.height}")
    
    # Images
    if ws_f._images:
        p(f"\n  ### 圖片: {len(ws_f._images)} 張")
        for img_idx, img in enumerate(ws_f._images):
            try:
                w = img.width
                h = img.height
            except:
                w, h = '?', '?'
            p(f"    圖片 {img_idx}: size=({w}x{h})")
    
    # Formulas
    p(f"\n  ### 公式:")
    formula_count = 0
    for row in ws_f.iter_rows(min_row=1, max_row=ws_f.max_row, max_col=ws_f.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
                # Also get the computed value
                data_cell = ws_d[cell.coordinate]
                computed = data_cell.value
                p(f"    {cell.coordinate}: {cell.value}")
                p(f"      → 計算值: {computed}")
    if formula_count == 0:
        p("    (無公式)")
    else:
        p(f"    共 {formula_count} 個公式")
    
    # Data content (first 100 rows, all columns)
    p(f"\n  ### 資料內容 (前 100 行):")
    for row_idx, row in enumerate(ws_d.iter_rows(min_row=1, max_row=min(ws_d.max_row, 100), max_col=ws_d.max_column, values_only=False), 1):
        row_data = []
        has_data = False
        for cell in row:
            val = cell.value
            if val is not None:
                has_data = True
                row_data.append(f"{cell.coordinate}={repr(val)}")
        if has_data:
            p(f"    Row {row_idx}: {' | '.join(row_data)}")
    
    # If more than 100 rows, show count
    if ws_d.max_row > 100:
        p(f"\n    ... (共 {ws_d.max_row} 行，僅顯示前 100 行)")
        # Show last few rows
        p(f"\n  ### 最後 10 行:")
        for row in ws_d.iter_rows(min_row=max(ws_d.max_row - 9, 101), max_row=ws_d.max_row, max_col=ws_d.max_column, values_only=False):
            row_data = []
            has_data = False
            for cell in row:
                val = cell.value
                if val is not None:
                    has_data = True
                    row_data.append(f"{cell.coordinate}={repr(val)}")
            if has_data:
                p(f"    Row {cell.row}: {' | '.join(row_data)}")

# Write output
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print(f"分析完成！輸出至 {OUTPUT_PATH}")
print(f"總行數: {len(out)}")
