#!/usr/bin/env python3
"""
Excel 公式透視系統 - 語義去重版
功能：提取邏輯上不同的公式，跨 DFA/FMEDA 比對相同概念公式
"""

import openpyxl
import json
import re
from pathlib import Path
from typing import Dict, List, Set, Tuple
from dataclasses import dataclass, asdict
from collections import defaultdict

@dataclass
class FormulaInfo:
    """公式資訊"""
    sheet_name: str
    cell_address: str
    original_formula: str
    normalized_formula: str  # 去除具體行列號後的公式
    formula_type: str  # CONDITIONAL, LOOKUP, AGGREGATE, MATH, EXPONENTIAL
    
    def to_dict(self):
        return asdict(self)

class SemanticFormulaExtractor:
    """語義去重公式提取器"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path, data_only=False)
        self.all_formulas: List[FormulaInfo] = []
        self.unique_patterns: Dict[str, List[FormulaInfo]] = {}  # 正規化公式 -> 所有出現位置
        
    def extract_all_formulas(self, sheets_to_scan: List[str] = None) -> List[FormulaInfo]:
        """提取所有公式"""
        print(f"開始從 {Path(self.excel_path).name} 提取公式...")
        
        sheets = sheets_to_scan if sheets_to_scan else self.wb.sheetnames
        
        for sheet_name in sheets:
            if sheet_name not in self.wb.sheetnames:
                print(f"  ⚠ Sheet '{sheet_name}' 不存在，跳過")
                continue
            
            ws = self.wb[sheet_name]
            print(f"  掃描 Sheet: {sheet_name} ({ws.max_row} 行 × {ws.max_column} 列)")
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        cell_addr = f"{sheet_name}!{cell.coordinate}"
                        
                        # 正規化公式
                        normalized = self._normalize_formula(formula)
                        formula_type = self._classify_formula(formula)
                        
                        formula_info = FormulaInfo(
                            sheet_name=sheet_name,
                            cell_address=cell_addr,
                            original_formula=formula,
                            normalized_formula=normalized,
                            formula_type=formula_type
                        )
                        
                        self.all_formulas.append(formula_info)
        
        print(f"✓ 共提取 {len(self.all_formulas)} 個公式")
        return self.all_formulas
    
    def deduplicate_by_semantics(self) -> Dict[str, List[FormulaInfo]]:
        """按語義去重（正規化公式相同 = 邏輯相同）"""
        print("\n開始語義去重...")
        
        for formula_info in self.all_formulas:
            normalized = formula_info.normalized_formula
            if normalized not in self.unique_patterns:
                self.unique_patterns[normalized] = []
            self.unique_patterns[normalized].append(formula_info)
        
        print(f"✓ 去重後共 {len(self.unique_patterns)} 個邏輯上不同的公式")
        
        # 按出現次數排序
        sorted_patterns = dict(sorted(
            self.unique_patterns.items(),
            key=lambda x: len(x[1]),
            reverse=True
        ))
        
        return sorted_patterns
    
    def _normalize_formula(self, formula: str) -> str:
        """正規化公式，去除具體行列號"""
        # 策略：將所有具體的行列號替換為通用符號
        
        # 1. 替換絕對引用中的行號（$A$1 -> $A$ROW）
        formula = re.sub(r'(\$?[A-Z]+)\$(\d+)', r'\1$ROW', formula)
        
        # 2. 替換相對引用中的行號（A1 -> A_ROW）
        formula = re.sub(r'(\$?[A-Z]+)(\d+)(?![0-9])', r'\1_ROW', formula)
        
        # 3. 替換範圍中的行號（$A$1:$A$100 -> $A$ROW:$A$ROW）
        formula = re.sub(r'(\$?[A-Z]+)\$(\d+):(\$?[A-Z]+)\$(\d+)', r'\1$ROW:\3$ROW', formula)
        
        return formula
    
    def _classify_formula(self, formula: str) -> str:
        """分類公式類型"""
        formula_upper = formula.upper()
        
        if 'EXP(' in formula_upper:
            return 'EXPONENTIAL'
        elif 'IF(' in formula_upper:
            return 'CONDITIONAL'
        elif any(func in formula_upper for func in ['VLOOKUP(', 'HLOOKUP(', 'LOOKUP(', 'INDEX(', 'MATCH(']):
            return 'LOOKUP'
        elif any(func in formula_upper for func in ['SUM(', 'AVERAGE(', 'COUNT(', 'MAX(', 'MIN(', 'RANK(']):
            return 'AGGREGATE'
        else:
            return 'MATH'
    
    def save_unique_formulas_to_markdown(self, output_path: str):
        """保存邏輯上不同的公式到 Markdown"""
        lines = []
        lines.append(f"# {Path(self.excel_path).name} - 邏輯上不同的公式\n\n")
        lines.append(f"**總公式數**: {len(self.all_formulas)}\n")
        lines.append(f"**邏輯上不同的公式**: {len(self.unique_patterns)}\n")
        lines.append(f"**平均重複次數**: {len(self.all_formulas) / len(self.unique_patterns):.1f}\n\n")
        
        # 按類型分組
        by_type = defaultdict(list)
        for normalized, infos in self.unique_patterns.items():
            ftype = infos[0].formula_type
            by_type[ftype].append((normalized, infos))
        
        for ftype in ['CONDITIONAL', 'LOOKUP', 'EXPONENTIAL', 'AGGREGATE', 'MATH']:
            if ftype in by_type:
                lines.append(f"## {ftype} 公式 ({len(by_type[ftype])} 個)\n\n")
                
                for idx, (normalized, infos) in enumerate(by_type[ftype], 1):
                    lines.append(f"### {ftype}-{idx}\n\n")
                    lines.append(f"**出現次數**: {len(infos)}\n\n")
                    lines.append(f"**原始公式** (第一次出現在 {infos[0].cell_address}):\n")
                    lines.append(f"```excel\n{infos[0].original_formula}\n```\n\n")
                    lines.append(f"**正規化公式**:\n")
                    lines.append(f"```\n{normalized}\n```\n\n")
                    lines.append(f"**出現位置**:\n")
                    for info in infos[:10]:  # 只顯示前 10 個
                        lines.append(f"- {info.cell_address}\n")
                    if len(infos) > 10:
                        lines.append(f"- ... 及其他 {len(infos) - 10} 個\n")
                    lines.append("\n---\n\n")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.writelines(lines)
        
        print(f"✓ 已保存 Markdown: {output_path}")
    
    def save_to_json(self, output_path: str):
        """保存為 JSON"""
        data = {
            'metadata': {
                'excel_file': self.excel_path,
                'total_formulas': len(self.all_formulas),
                'unique_patterns': len(self.unique_patterns),
                'average_repetition': len(self.all_formulas) / len(self.unique_patterns) if self.unique_patterns else 0
            },
            'patterns': {}
        }
        
        for idx, (normalized, infos) in enumerate(self.unique_patterns.items(), 1):
            pattern_key = f"P{idx:04d}"
            data['patterns'][pattern_key] = {
                'normalized_formula': normalized,
                'formula_type': infos[0].formula_type,
                'first_occurrence': infos[0].cell_address,
                'original_formula': infos[0].original_formula,
                'occurrence_count': len(infos),
                'all_occurrences': [info.cell_address for info in infos]
            }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"✓ 已保存 JSON: {output_path}")

def compare_dfa_fmeda_formulas():
    """比對 DFA 和 FMEDA 的相同概念公式"""
    print("\n" + "=" * 80)
    print("比對 DFA 和 FMEDA 的相同概念公式")
    print("=" * 80)
    
    # 提取 DFA 公式
    dfa_extractor = SemanticFormulaExtractor(
        '/home/ubuntu/upload/RD-03-009-01.ADFA報告DFAReport.xlsx'
    )
    dfa_extractor.extract_all_formulas()
    dfa_patterns = dfa_extractor.deduplicate_by_semantics()
    
    # 提取 FMEDA 公式
    fmeda_extractor = SemanticFormulaExtractor(
        '/home/ubuntu/upload/RD-03-008-01FMEDAReport.xlsx'
    )
    fmeda_extractor.extract_all_formulas()
    fmeda_patterns = fmeda_extractor.deduplicate_by_semantics()
    
    # 找相同的正規化公式
    dfa_normalized_set = set(dfa_patterns.keys())
    fmeda_normalized_set = set(fmeda_patterns.keys())
    
    common_patterns = dfa_normalized_set & fmeda_normalized_set
    dfa_only = dfa_normalized_set - fmeda_normalized_set
    fmeda_only = fmeda_normalized_set - dfa_normalized_set
    
    print(f"\n共同公式: {len(common_patterns)}")
    print(f"DFA 獨有: {len(dfa_only)}")
    print(f"FMEDA 獨有: {len(fmeda_only)}")
    
    # 保存比對結果
    comparison = {
        'dfa_total': len(dfa_extractor.all_formulas),
        'dfa_unique': len(dfa_patterns),
        'fmeda_total': len(fmeda_extractor.all_formulas),
        'fmeda_unique': len(fmeda_patterns),
        'common_patterns': len(common_patterns),
        'dfa_only': len(dfa_only),
        'fmeda_only': len(fmeda_only),
        'common_formulas': list(common_patterns)[:20]  # 只保存前 20 個
    }
    
    with open('/home/ubuntu/dfa_fmeda_formula_comparison.json', 'w', encoding='utf-8') as f:
        json.dump(comparison, f, ensure_ascii=False, indent=2)
    
    print(f"✓ 已保存比對結果: /home/ubuntu/dfa_fmeda_formula_comparison.json")
    
    return dfa_extractor, fmeda_extractor

def main():
    """主程式"""
    print("=" * 80)
    print("Excel 公式透視系統 - 語義去重版")
    print("=" * 80)
    
    # 提取 FMEDA 公式（關鍵 sheets）
    print("\n[1/2] 提取 FMEDA 公式...")
    fmeda_extractor = SemanticFormulaExtractor(
        '/home/ubuntu/upload/RD-03-008-01FMEDAReport.xlsx'
    )
    fmeda_extractor.extract_all_formulas([
        'BlockList',
        'FailureRateCalc',
        'FailureRateCalcIC',
        'FMEDA'
    ])
    fmeda_extractor.deduplicate_by_semantics()
    fmeda_extractor.save_unique_formulas_to_markdown('/home/ubuntu/FMEDA_unique_formulas.md')
    fmeda_extractor.save_to_json('/home/ubuntu/FMEDA_unique_formulas.json')
    
    # 比對 DFA 和 FMEDA
    print("\n[2/2] 比對 DFA 和 FMEDA...")
    dfa_extractor, fmeda_extractor_full = compare_dfa_fmeda_formulas()
    
    print("\n" + "=" * 80)
    print("完成！")
    print("=" * 80)

if __name__ == '__main__':
    main()
