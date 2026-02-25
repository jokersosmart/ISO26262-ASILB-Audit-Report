#!/usr/bin/env python3
"""
FMEDA Excel to Reversible Markdown Converter
將 RD-03-008-01FMEDAReport.xlsx 轉換為可逆向還原的 Markdown 格式
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import warnings
import os

warnings.filterwarnings('ignore')

INPUT_FILE = '/home/ubuntu/upload/RD-03-008-01FMEDAReport.xlsx'
OUTPUT_FILE = '/home/ubuntu/FMEDA_Analysis/markdown/FMEDA_Report_Reversible.md'

os.makedirs('/home/ubuntu/FMEDA_Analysis/markdown', exist_ok=True)

def safe_str(val):
    if val is None:
        return ''
    return str(val).replace('|', '\\|').replace('\n', ' ')

def color_to_hex(color):
    try:
        if color and color.rgb and color.rgb != '00000000':
            rgb = str(color.rgb)
            if len(rgb) == 8:
                return '#' + rgb[2:]
            elif len(rgb) == 6:
                return '#' + rgb
    except:
        pass
    return None

def get_formula(ws_formula, row, col):
    """Get formula from formula-mode worksheet"""
    try:
        cell = ws_formula.cell(row=row, column=col)
        if cell.value and str(cell.value).startswith('='):
            return str(cell.value)
    except:
        pass
    return None

print("Loading workbook (data_only=True)...")
wb_data = openpyxl.load_workbook(INPUT_FILE, data_only=True)
print("Loading workbook (data_only=False for formulas)...")
wb_formula = openpyxl.load_workbook(INPUT_FILE, data_only=False)

lines = []
lines.append('# RD-03-008-01 FMEDA Report -- Reversible Markdown')
lines.append('')
lines.append('<!-- METADATA')
lines.append(json.dumps({
    'source_file': 'RD-03-008-01FMEDAReport.xlsx',
    'sheet_count': len(wb_data.sheetnames),
    'sheet_names': wb_data.sheetnames,
    'format_version': '2.0',
    'reversible': True
}, indent=2, ensure_ascii=False))
lines.append('-->')
lines.append('')

# Table of Contents
lines.append('## Table of Contents')
lines.append('')
for i, name in enumerate(wb_data.sheetnames, 1):
    anchor = name.replace(' ', '-').replace('.', '').replace('(', '').replace(')', '')
    lines.append(f'{i}. [{name}](#{anchor})')
lines.append('')
lines.append('---')
lines.append('')

# Process each sheet
for sheet_idx, sheet_name in enumerate(wb_data.sheetnames):
    print(f"Processing sheet {sheet_idx+1}/{len(wb_data.sheetnames)}: {sheet_name}")
    
    ws_data = wb_data[sheet_name]
    ws_form = wb_formula[sheet_name]
    
    # Sheet header
    lines.append(f'## {sheet_name}')
    lines.append('')
    
    # Sheet metadata
    visibility = 'visible'
    if ws_data.sheet_state == 'hidden':
        visibility = 'hidden'
    elif ws_data.sheet_state == 'veryHidden':
        visibility = 'veryHidden'
    
    # Merged cells
    merged = []
    for mc in ws_data.merged_cells.ranges:
        merged.append(str(mc))
    
    meta = {
        'sheet_index': sheet_idx,
        'visibility': visibility,
        'max_row': ws_data.max_row,
        'max_col': ws_data.max_column,
        'merged_cells': merged[:50],  # Limit for readability
        'total_merged': len(merged)
    }
    
    # Column widths
    col_widths = {}
    for col_dim in ws_data.column_dimensions:
        if ws_data.column_dimensions[col_dim].width:
            col_widths[col_dim] = round(ws_data.column_dimensions[col_dim].width, 2)
    if col_widths:
        meta['column_widths_sample'] = dict(list(col_widths.items())[:20])
    
    lines.append('<!-- SHEET_META')
    lines.append(json.dumps(meta, indent=2, ensure_ascii=False))
    lines.append('-->')
    lines.append('')
    
    # Determine data range
    max_row = min(ws_data.max_row or 1, 2000)  # Cap at 2000 rows for large sheets
    max_col = min(ws_data.max_column or 1, 100)  # Cap at 100 cols
    
    # For very large sheets (like FMEDA main table), limit output
    is_large = (ws_data.max_row or 0) > 500 and (ws_data.max_column or 0) > 50
    
    if is_large:
        lines.append(f'> **Note**: This sheet has {ws_data.max_row} rows x {ws_data.max_column} columns.')
        lines.append(f'> Showing first 100 rows. Full data preserved in SHEET_DATA block below.')
        lines.append('')
        max_row = min(100, ws_data.max_row or 1)
        max_col = min(ws_data.max_column or 1, 100)
    
    # Find actual data bounds (skip empty rows/cols)
    actual_max_row = 0
    actual_max_col = 0
    for row in ws_data.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                actual_max_row = max(actual_max_row, cell.row)
                actual_max_col = max(actual_max_col, cell.column)
    
    if actual_max_row == 0:
        lines.append('*Empty sheet*')
        lines.append('')
        lines.append('---')
        lines.append('')
        continue
    
    actual_max_row = min(actual_max_row, max_row)
    actual_max_col = min(actual_max_col, max_col)
    
    # Collect formulas for this sheet
    formula_cells = {}
    for row in range(1, actual_max_row + 1):
        for col in range(1, actual_max_col + 1):
            f = get_formula(ws_form, row, col)
            if f:
                coord = f"{get_column_letter(col)}{row}"
                formula_cells[coord] = f
    
    if formula_cells:
        # Show formula summary
        lines.append(f'### Formulas ({len(formula_cells)} total)')
        lines.append('')
        lines.append('<!-- FORMULAS')
        # For large formula sets, show sample
        sample_size = min(50, len(formula_cells))
        sample = dict(list(formula_cells.items())[:sample_size])
        lines.append(json.dumps(sample, indent=2, ensure_ascii=False))
        if len(formula_cells) > sample_size:
            lines.append(f'// ... and {len(formula_cells) - sample_size} more formulas')
        lines.append('-->')
        lines.append('')
        
        # Show key formulas in readable format
        lines.append('**Key Formula Samples:**')
        lines.append('')
        lines.append('| Cell | Formula | Computed Value |')
        lines.append('|------|---------|---------------|')
        shown = 0
        for coord, formula in formula_cells.items():
            if shown >= 20:
                break
            # Get computed value
            col_letter = ''.join(c for c in coord if c.isalpha())
            row_num = int(''.join(c for c in coord if c.isdigit()))
            col_num = openpyxl.utils.column_index_from_string(col_letter)
            val = ws_data.cell(row=row_num, column=col_num).value
            lines.append(f'| `{coord}` | `{safe_str(formula)[:80]}` | `{safe_str(val)[:40]}` |')
            shown += 1
        lines.append('')
    
    # Data table
    lines.append('### Data')
    lines.append('')
    
    # Build header row
    header = []
    for col in range(1, actual_max_col + 1):
        header.append(get_column_letter(col))
    
    lines.append('| ' + ' | '.join(header) + ' |')
    lines.append('| ' + ' | '.join(['---'] * len(header)) + ' |')
    
    # Data rows
    for row in range(1, actual_max_row + 1):
        row_data = []
        for col in range(1, actual_max_col + 1):
            cell = ws_data.cell(row=row, column=col)
            val = safe_str(cell.value)
            if len(val) > 60:
                val = val[:57] + '...'
            row_data.append(val)
        lines.append('| ' + ' | '.join(row_data) + ' |')
    
    lines.append('')
    
    # Full formula dump for large sheets (in hidden block)
    if len(formula_cells) > 50:
        lines.append('<!-- FULL_FORMULAS')
        # Dump in batches
        batch = {}
        count = 0
        for coord, formula in formula_cells.items():
            batch[coord] = formula
            count += 1
            if count % 500 == 0:
                lines.append(json.dumps(batch, ensure_ascii=False))
                batch = {}
        if batch:
            lines.append(json.dumps(batch, ensure_ascii=False))
        lines.append('-->')
        lines.append('')
    
    lines.append('---')
    lines.append('')

# Write output
print(f"Writing output to {OUTPUT_FILE}...")
with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

line_count = len(lines)
print(f"Done! Output: {OUTPUT_FILE} ({line_count} lines)")
